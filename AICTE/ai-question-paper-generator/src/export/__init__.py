"""
Spreadsheet Generator Module

Generates formatted Excel and CSV spreadsheets from selected questions.
Supports various output formats and customizable styling.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any, Optional
import logging
from pathlib import Path
import json


class SpreadsheetGenerator:
    """Generates formatted spreadsheets from question data"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.default_columns = [
            'id', 'question', 'topic', 'difficulty', 'type', 'keywords', 'answer'
        ]
        
    def generate_spreadsheet(self, questions: List[Dict[str, Any]], 
                           output_path: str, 
                           format_type: str = 'excel',
                           columns: Optional[List[str]] = None,
                           style: Optional[Dict[str, Any]] = None) -> bool:
        """
        Generate spreadsheet from questions
        
        Args:
            questions: List of question dictionaries
            output_path: Path to output file
            format_type: 'excel' or 'csv'
            columns: List of columns to include
            style: Styling options for Excel format
            
        Returns:
            Success status
        """
        try:
            output_path = Path(output_path)
            
            # Determine format from extension if not specified
            if format_type == 'excel' or output_path.suffix.lower() in ['.xlsx', '.xls']:
                return self._generate_excel(questions, output_path, columns, style)
            else:
                return self._generate_csv(questions, output_path, columns)
                
        except Exception as e:
            self.logger.error(f"Error generating spreadsheet: {str(e)}")
            return False
    
    def _generate_excel(self, questions: List[Dict[str, Any]], 
                       output_path: Path,
                       columns: Optional[List[str]] = None,
                       style: Optional[Dict[str, Any]] = None) -> bool:
        """Generate Excel spreadsheet"""
        try:
            # Prepare data
            df = self._prepare_dataframe(questions, columns)
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Selected Questions"
            
            # Write data
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply styling
            if style is not None:
                self._apply_excel_styling(ws, style)
            else:
                self._apply_default_styling(ws)
            
            # Adjust column widths
            self._adjust_column_widths(ws)
            
            # Add metadata sheet
            self._add_metadata_sheet(wb, questions)
            
            # Save workbook
            wb.save(output_path)
            
            self.logger.info(f"Excel spreadsheet generated: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error generating Excel file: {str(e)}")
            return False
    
    def _generate_csv(self, questions: List[Dict[str, Any]], 
                     output_path: Path,
                     columns: Optional[List[str]] = None) -> bool:
        """Generate CSV spreadsheet"""
        try:
            # Prepare data
            df = self._prepare_dataframe(questions, columns)
            
            # Save to CSV
            df.to_csv(output_path, index=False, encoding='utf-8')
            
            self.logger.info(f"CSV spreadsheet generated: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error generating CSV file: {str(e)}")
            return False
    
    def _prepare_dataframe(self, questions: List[Dict[str, Any]], 
                          columns: Optional[List[str]] = None) -> pd.DataFrame:
        """Prepare DataFrame from questions"""
        if columns is None:
            columns = self.default_columns
        
        # Prepare data rows
        data_rows = []
        for question in questions:
            row = {}
            for col in columns:
                value = question.get(col, '')
                
                # Handle list values
                if isinstance(value, list):
                    value = ', '.join(str(v) for v in value)
                
                row[col] = value
            
            data_rows.append(row)
        
        return pd.DataFrame(data_rows)
    
    def _apply_default_styling(self, ws):
        """Apply default styling to Excel worksheet"""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data styling
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Apply data styling
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def _apply_excel_styling(self, ws, style: Dict[str, Any]):
        """Apply custom styling to Excel worksheet"""
        # Header styling
        if 'header' in style:
            header_style = style['header']
            for cell in ws[1]:
                if 'font' in header_style:
                    cell.font = Font(**header_style['font'])
                if 'fill' in header_style:
                    cell.fill = PatternFill(**header_style['fill'])
                if 'alignment' in header_style:
                    cell.alignment = Alignment(**header_style['alignment'])
        
        # Data styling
        if 'data' in style:
            data_style = style['data']
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if 'font' in data_style:
                        cell.font = Font(**data_style['font'])
                    if 'alignment' in data_style:
                        cell.alignment = Alignment(**data_style['alignment'])
        
        # Border styling
        if 'border' in style:
            border = Border(**style['border'])
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
    
    def _adjust_column_widths(self, ws):
        """Adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set width with reasonable limits
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_metadata_sheet(self, wb, questions: List[Dict[str, Any]]):
        """Add metadata sheet with statistics"""
        metadata_ws = wb.create_sheet("Metadata")
        
        # Statistics
        stats = self._calculate_statistics(questions)
        
        # Write statistics
        row = 1
        metadata_ws.cell(row=row, column=1, value="Question Bank Statistics")
        metadata_ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        for key, value in stats.items():
            metadata_ws.cell(row=row, column=1, value=key)
            metadata_ws.cell(row=row, column=2, value=str(value))
            row += 1
        
        # Adjust column widths
        self._adjust_column_widths(metadata_ws)
    
    def _calculate_statistics(self, questions: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate statistics for metadata"""
        from collections import Counter
        
        stats = {
            'Total Questions': len(questions),
            'Topics': len(set(q.get('topic', '') for q in questions)),
            'Difficulties': len(set(q.get('difficulty', '') for q in questions)),
            'Types': len(set(q.get('type', '') for q in questions))
        }
        
        # Topic distribution
        topic_counts = Counter(q.get('topic', 'unknown') for q in questions)
        stats['Most Common Topic'] = topic_counts.most_common(1)[0][0] if topic_counts else 'N/A'
        
        # Difficulty distribution
        difficulty_counts = Counter(q.get('difficulty', 'unknown') for q in questions)
        stats['Most Common Difficulty'] = difficulty_counts.most_common(1)[0][0] if difficulty_counts else 'N/A'
        
        # Average question length
        lengths = [len(q.get('question', '')) for q in questions]
        stats['Average Question Length'] = sum(lengths) / len(lengths) if lengths else 0
        
        return stats
    
    def generate_multiple_sheets(self, questions_by_category: Dict[str, List[Dict[str, Any]]], 
                               output_path: str) -> bool:
        """Generate Excel file with multiple sheets by category"""
        try:
            output_path = Path(output_path)
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheet for each category
            for category, questions in questions_by_category.items():
                ws = wb.create_sheet(category)
                
                # Prepare data
                df = self._prepare_dataframe(questions)
                
                # Write data
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                # Apply styling
                self._apply_default_styling(ws)
                self._adjust_column_widths(ws)
            
            # Add summary sheet
            self._add_summary_sheet(wb, questions_by_category)
            
            # Save workbook
            wb.save(output_path)
            
            self.logger.info(f"Multi-sheet Excel file generated: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error generating multi-sheet Excel file: {str(e)}")
            return False
    
    def _add_summary_sheet(self, wb, questions_by_category: Dict[str, List[Dict[str, Any]]]):
        """Add summary sheet with category statistics"""
        summary_ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
        
        # Headers
        summary_ws.cell(row=1, column=1, value="Category")
        summary_ws.cell(row=1, column=2, value="Question Count")
        summary_ws.cell(row=1, column=3, value="Average Length")
        
        # Apply header styling
        for cell in summary_ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data rows
        row = 2
        for category, questions in questions_by_category.items():
            summary_ws.cell(row=row, column=1, value=category)
            summary_ws.cell(row=row, column=2, value=len(questions))
            
            if questions:
                avg_length = sum(len(q.get('question', '')) for q in questions) / len(questions)
                summary_ws.cell(row=row, column=3, value=round(avg_length, 1))
            else:
                summary_ws.cell(row=row, column=3, value=0)
            
            row += 1
        
        # Adjust column widths
        self._adjust_column_widths(summary_ws)
    
    def export_to_json(self, questions: List[Dict[str, Any]], 
                      output_path: str) -> bool:
        """Export questions to JSON format"""
        try:
            output_path = Path(output_path)
            
            # Prepare data
            export_data = {
                'metadata': {
                    'total_questions': len(questions),
                    'export_timestamp': pd.Timestamp.now().isoformat()
                },
                'questions': questions
            }
            
            # Save to JSON
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            self.logger.info(f"JSON export completed: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error exporting to JSON: {str(e)}")
            return False
